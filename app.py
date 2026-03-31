import streamlit as st
import pandas as pd
import re
import io
import zipfile
import os
from openpyxl.styles import PatternFill

# Page Configuration
st.set_page_config(page_title="Robot Signal Extractor", page_icon="🤖", layout="wide")

# Custom CSS for Premium Look
st.markdown("""
    <style>
    .main {
        background-color: #0e1117;
    }
    .stButton>button {
        width: 100%;
        border-radius: 5px;
        height: 3em;
        background-color: #007bff;
        color: white;
        font-weight: bold;
    }
    .stDownloadButton>button {
        width: 100%;
        border-radius: 5px;
        height: 3em;
        background-color: #28a745;
        color: white;
        font-weight: bold;
    }
    .css-1offfwp {
        background-color: rgba(255, 255, 255, 0.05);
        border-radius: 10px;
        padding: 20px;
        backdrop-filter: blur(10px);
    }
    </style>
    """, unsafe_allow_html=True)

# Parsing Logic (Adapted for Zip)
def extract_signals_from_text(content):
    di_entries = []
    do_entries = []
    di_pattern = re.compile(r'DI\[(\d+)(?::([^\]]*))?\](?:\s*(?:=|==|<>|:)\s*(ON|OFF))?', re.IGNORECASE)
    do_pattern = re.compile(r'DO\[(\d+)(?::([^\]]*))?\](?:\s*(?:=|==|<>|:)\s*(ON|OFF))?', re.IGNORECASE)
    
    for line in content.splitlines():
        for match in di_pattern.finditer(line):
            index = int(match.group(1))
            comment = (match.group(2) or "").strip()
            state = (match.group(3) or "").upper()
            di_entries.append((index, comment, state))
        for match in do_pattern.finditer(line):
            index = int(match.group(1))
            comment = (match.group(2) or "").strip()
            state = (match.group(3) or "").upper()
            do_entries.append((index, comment, state))
    return di_entries, do_entries

def process_zip_backup(uploaded_file):
    with zipfile.ZipFile(uploaded_file, 'r') as z:
        file_list = z.namelist()
        
        # Identify robot folders (those containing an 'LS' folder or .ls files)
        # We group by the folder name before /LS/
        robot_data = {} # robot_name -> list of signal entries
        
        for name in file_list:
            if name.lower().endswith('.ls'):
                # Extract robot name from path
                # Pattern: common/path/ROBOT_NAME/LS/file.ls or ROBOT_NAME/file.ls
                path_parts = name.split('/')
                robot_name = "Unknown"
                
                # Check for /LS/ folder
                if 'LS' in [p.upper() for p in path_parts]:
                    ls_idx = [i for i, p in enumerate(path_parts) if p.upper() == 'LS'][0]
                    if ls_idx > 0:
                        robot_name = path_parts[ls_idx - 1]
                    else:
                        robot_name = "Main"
                else:
                    # Use immediate parent folder as robot name
                    if len(path_parts) > 1:
                        robot_name = path_parts[-2]
                    else:
                        robot_name = "Root"

                # Read and parse
                try:
                    with z.open(name) as f:
                        content = f.read().decode('latin-1', errors='replace')
                        di_list, do_list = extract_signals_from_text(content)
                        
                        if robot_name not in robot_data:
                            robot_data[robot_name] = {'DI': {}, 'DO': {}}
                        
                        for index, comment, state in di_list:
                            if not comment and not state: continue
                            curr_comment, curr_state = robot_data[robot_name]['DI'].get(index, ("", ""))
                            new_comment = comment or curr_comment
                            new_state = curr_state
                            prefix_match = re.match(r'^(ON|OFF)\s*:', new_comment, re.IGNORECASE)
                            if prefix_match:
                                new_state = prefix_match.group(1).upper()
                                new_comment = re.sub(r'^(ON|OFF)\s*:\s*', '', new_comment, flags=re.IGNORECASE)
                            elif state and not new_state: new_state = state
                            if not curr_comment or new_state or len(new_comment) > len(curr_comment):
                                robot_data[robot_name]['DI'][index] = (new_comment, new_state)
                                
                        for index, comment, state in do_list:
                            if not comment and not state: continue
                            curr_comment, curr_state = robot_data[robot_name]['DO'].get(index, ("", ""))
                            new_comment = comment or curr_comment
                            new_state = curr_state
                            prefix_match = re.match(r'^(ON|OFF)\s*:', new_comment, re.IGNORECASE)
                            if prefix_match:
                                new_state = prefix_match.group(1).upper()
                                new_comment = re.sub(r'^(ON|OFF)\s*:\s*', '', new_comment, flags=re.IGNORECASE)
                            elif state and not new_state: new_state = state
                            if not curr_comment or new_state or len(new_comment) > len(curr_comment):
                                robot_data[robot_name]['DO'][index] = (new_comment, new_state)
                except Exception as e:
                    st.warning(f"Error reading {name}: {e}")

        # Generate Excel in memory
        output = io.BytesIO()
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for robot_name, signals in robot_data.items():
                all_di = signals['DI']
                all_do = signals['DO']
                has_state = any(s for _, s in all_di.values()) or any(s for _, s in all_do.values())
                
                di_rows = []
                do_rows = []
                for i in sorted(all_di.keys()):
                    comment, state = all_di[i]
                    if has_state: di_rows.append([f"DI {i}", state, comment])
                    else: di_rows.append([f"DI {i}", comment])
                for i in sorted(all_do.keys()):
                    comment, state = all_do[i]
                    if has_state: do_rows.append([f"DO {i}", state, comment])
                    else: do_rows.append([f"DO {i}", comment])
                
                if not di_rows and not do_rows: continue
                
                # Create DFs
                if has_state:
                    df_di = pd.DataFrame(di_rows, columns=['DI', 'State', 'Comment'])
                    df_do = pd.DataFrame(do_rows, columns=['DO', 'State', 'Comment'])
                else:
                    df_di = pd.DataFrame(di_rows, columns=['DI', 'Comment'])
                    df_do = pd.DataFrame(do_rows, columns=['DO', 'Comment'])
                    
                max_rows = max(len(df_di), len(df_do))
                col_space = pd.DataFrame([''] * max_rows, columns=[' '])
                df_do_padded = df_do.reindex(range(max_rows)).fillna('')
                df_di_padded = df_di.reindex(range(max_rows)).fillna('')
                
                final_df = pd.concat([df_do_padded, col_space, df_di_padded], axis=1)
                sheet_name = str(robot_name)[:31]
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                ws = writer.sheets[sheet_name]
                if has_state:
                    ws.column_dimensions['A'].width = 11
                    ws.column_dimensions['B'].width = 8
                    ws.column_dimensions['C'].width = 30
                    ws.column_dimensions['D'].width = 5
                    ws.column_dimensions['E'].width = 11
                    ws.column_dimensions['F'].width = 8
                    ws.column_dimensions['G'].width = 30
                    for row_idx in range(2, max_rows + 2):
                        # Col B (DO State) and Col F (DI State)
                        sv_do = ws.cell(row=row_idx, column=2).value
                        if sv_do == 'ON': ws.cell(row=row_idx, column=2).fill = green_fill
                        elif sv_do == 'OFF': ws.cell(row=row_idx, column=2).fill = red_fill
                        sv_di = ws.cell(row=row_idx, column=6).value
                        if sv_di == 'ON': ws.cell(row=row_idx, column=6).fill = green_fill
                        elif sv_di == 'OFF': ws.cell(row=row_idx, column=6).fill = red_fill
                else:
                    ws.column_dimensions['A'].width = 13
                    ws.column_dimensions['B'].width = 27
                    ws.column_dimensions['C'].width = 13
                    ws.column_dimensions['D'].width = 13
                    ws.column_dimensions['E'].width = 21

        return output.getvalue(), list(robot_data.keys())

# Main App Layout
st.title("🤖 Robot Signal Extractor")
st.markdown("### Professional Industrial I/O Signal Analysis Tool")
st.write("Upload a `.zip` file of your robot backup to automatically extract DI/DO signals into a formatted Excel table.")

with st.container():
    uploaded_file = st.file_uploader("Drop your robot backup ZIP here", type="zip")

if uploaded_file:
    # Get original filename minus extension
    base_name = os.path.splitext(uploaded_file.name)[0]
    
    with st.spinner("Processing robot signals..."):
        excel_data, robots_found = process_zip_backup(uploaded_file)
        
    if excel_data:
        st.success(f"✅ Successfully processed {len(robots_found)} robot(s)!")
        
        # Display Preview of Robots Found
        st.write("##### Robots Identified:")
        st.info(", ".join(robots_found))
        
        # Download Button
        st.download_button(
            label="📥 Download Excel Signal Table",
            data=excel_data,
            file_name=f"{base_name}_Signals.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("No valid .ls files found in the ZIP. Please check your backup structure.")

# Footer
st.markdown("---")
st.caption("Powered by Advanced Agentic Automation | Streamlit Deployment Ready")
